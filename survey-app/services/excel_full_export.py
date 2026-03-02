"""
Full database export to Excel — one sheet per table.

Sheets produced:
  1. Клиенты          — all Client rows
  2. Сотрудники       — all Employee rows
  3. Опросы           — all Survey rows (full detail)
  4. Жалобы (сотрудники) — surveys where complaint_employee=True
  5. Жалобы (условия)    — surveys where complaint_conditions=True
"""
import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import Client, Employee, Survey

log = logging.getLogger(__name__)

# ── Header style ─────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes        = "A2"
    ws.row_dimensions[1].height = 30


def _auto_width(ws) -> None:
    """Fit column widths to content (max 55 chars)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


def _d(val: date | None) -> str:
    return val.strftime("%d.%m.%Y") if val else ""


def _e(val) -> str:
    """Enum → str."""
    if val is None:
        return ""
    return val.value if hasattr(val, "value") else str(val)


def _b(val: bool) -> str:
    return "Да" if val else "Нет"


# ── Main export function ───────────────────────────────────────────────────────

def export_full_excel(session: Session, path: str | Path) -> dict[str, int]:
    """
    Export all tables to an Excel workbook.

    Returns dict ``{sheet_name: row_count}`` (not counting the header row).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)           # remove the default empty sheet
    counts: dict[str, int] = {}

    # ── 1. Клиенты ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Клиенты")
    _write_header(ws, [
        "ID", "Имя ребёнка", "Имя родителя",
        "Дата начала", "Статус", "Специалисты",
    ])
    clients = session.query(Client).order_by(Client.child_name).all()
    for c in clients:
        specialists = ", ".join(e.full_name for e in c.specialists)
        ws.append([
            c.id,
            c.child_name,
            c.parent_name or "",
            _d(c.start_date),
            _e(c.status),
            specialists,
        ])
    _auto_width(ws)
    counts["Клиенты"] = len(clients)
    log.debug("Sheet 'Клиенты': %d rows", len(clients))

    # ── 2. Сотрудники ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Сотрудники")
    _write_header(ws, ["ID", "ФИО", "Должность", "Статус"])
    employees = session.query(Employee).order_by(Employee.full_name).all()
    for e in employees:
        ws.append([e.id, e.full_name, _e(e.position), _e(e.status)])
    _auto_width(ws)
    counts["Сотрудники"] = len(employees)
    log.debug("Sheet 'Сотрудники': %d rows", len(employees))

    # ── 3. Опросы ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Опросы")
    _write_header(ws, [
        "ID", "Клиент (ребёнок)", "Дата контакта", "Тип контакта", "Провёл",
        "Удовлетворённость", "Непонимание",
        "Жалоба на сотрудника", "Текст (сотрудник)",
        "Жалоба на условия",   "Текст (условия)",
        "Статус ситуации", "Результат урегулирования", "Причина неурегулирования",
        "Специалисты (снимок)", "Сотрудники жалобы",
        "Комментарий",
    ])
    surveys = (
        session.query(Survey)
        .join(Client)
        .order_by(Survey.contact_date.asc().nullslast(), Survey.id)
        .all()
    )
    for s in surveys:
        snap      = ", ".join(e.full_name for e in s.specialists_snapshot)
        comp_emps = ", ".join(e.full_name for e in s.complaint_employees)
        ws.append([
            s.id,
            s.client.child_name,
            _d(s.contact_date),
            _e(s.contact_type),
            s.conducted_by or "",
            _e(s.satisfaction),
            _e(s.misunderstanding),
            _b(s.complaint_employee),
            s.complaint_employee_text or "",
            _b(s.complaint_conditions),
            s.complaint_conditions_text or "",
            _e(s.situation_status),
            s.resolution_result or "",
            s.non_resolution_reason or "",
            snap,
            comp_emps,
            s.comment_text or "",
        ])
    _auto_width(ws)
    counts["Опросы"] = len(surveys)
    log.debug("Sheet 'Опросы': %d rows", len(surveys))

    # ── 4. Жалобы на сотрудников ────────────────────────────────────────────
    ws = wb.create_sheet("Жалобы на сотрудников")
    _write_header(ws, [
        "ID опроса", "Клиент", "Дата",
        "Сотрудники (жалоба)", "Текст жалобы",
        "Статус ситуации", "Результат",
    ])
    emp_complaints = [s for s in surveys if s.complaint_employee]
    for s in emp_complaints:
        comp_emps = ", ".join(e.full_name for e in s.complaint_employees)
        ws.append([
            s.id,
            s.client.child_name,
            _d(s.contact_date),
            comp_emps,
            s.complaint_employee_text or "",
            _e(s.situation_status),
            s.resolution_result or "",
        ])
    _auto_width(ws)
    counts["Жалобы на сотрудников"] = len(emp_complaints)
    log.debug("Sheet 'Жалобы на сотрудников': %d rows", len(emp_complaints))

    # ── 5. Жалобы на условия ─────────────────────────────────────────────────
    ws = wb.create_sheet("Жалобы на условия")
    _write_header(ws, [
        "ID опроса", "Клиент", "Дата",
        "Текст жалобы",
        "Статус ситуации", "Результат",
    ])
    cond_complaints = [s for s in surveys if s.complaint_conditions]
    for s in cond_complaints:
        ws.append([
            s.id,
            s.client.child_name,
            _d(s.contact_date),
            s.complaint_conditions_text or "",
            _e(s.situation_status),
            s.resolution_result or "",
        ])
    _auto_width(ws)
    counts["Жалобы на условия"] = len(cond_complaints)
    log.debug("Sheet 'Жалобы на условия': %d rows", len(cond_complaints))

    wb.save(path)
    log.info("Full Excel export saved → %s | rows: %s", path, counts)
    return counts
